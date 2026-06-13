"""各種 Exporter のユニットテスト。"""

from pathlib import Path
import pytest

from app.core.exporter import ExporterFactory, TxtExporter, DocxExporter, XlsxExporter, PdfExporter
from app.exceptions import ExportError
from app.models.template_model import FieldMapping, Template


def test_txt_exporter(tmp_path):
    output_file = tmp_path / "result.txt"
    mapped_data = {"顧客名": "テスト太郎", "金額": 150000}
    
    template = Template(
        name="テストテンプレート",
        output_format="txt",
        output_filename_pattern="test.txt",
        fields=[
            FieldMapping(source_key="name", output_label="顧客名", target_position=""),
            FieldMapping(source_key="price", output_label="金額", target_position=""),
        ]
    )
    
    exporter = ExporterFactory.create("txt")
    exporter.export(mapped_data, template, output_file)
    
    assert output_file.exists()
    content = output_file.read_text(encoding="utf-8")
    assert "顧客名: テスト太郎" in content
    assert "金額: 150000" in content


def test_xlsx_exporter_new_file(tmp_path):
    output_file = tmp_path / "result.xlsx"
    mapped_data = {"顧客名": "テスト太郎", "金額": 150000}
    
    template = Template(
        name="テストテンプレート",
        output_format="xlsx",
        output_filename_pattern="test.xlsx",
        fields=[
            FieldMapping(source_key="name", output_label="顧客名", target_position="A2"),
            FieldMapping(source_key="price", output_label="金額", target_position="B2"),
        ]
    )
    
    # openpyxl がインストールされているか確認
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl がインストールされていません")
        
    exporter = ExporterFactory.create("xlsx")
    exporter.export(mapped_data, template, output_file)
    
    assert output_file.exists()
    
    # 読み込んで確認
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    # 新規作成時はテンプレートファイルなしのため、デフォルト出力（A2に顧客名、B2に金額の値）になる。
    # テンプレートなしのデフォルト出力は:
    # 行2に: col1 = "顧客名", col2 = "テスト太郎"
    # 行3に: col3 = "金額", col2 = 150000
    # のようになる設計になっていることを確認。
    row2_label = ws.cell(row=2, column=1).value
    row2_val = ws.cell(row=2, column=2).value
    assert row2_label == "顧客名"
    assert row2_val == "テスト太郎"


def test_docx_exporter_new_file(tmp_path):
    output_file = tmp_path / "result.docx"
    mapped_data = {"顧客名": "テスト太郎"}
    
    template = Template(
        name="テストテンプレート",
        output_format="docx",
        output_filename_pattern="test.docx",
        fields=[
            FieldMapping(source_key="name", output_label="顧客名", target_position=""),
        ]
    )
    
    try:
        import docx
    except ImportError:
        pytest.skip("python-docx がインストールされていません")
        
    exporter = ExporterFactory.create("docx")
    exporter.export(mapped_data, template, output_file)
    
    assert output_file.exists()


def test_pdf_exporter(tmp_path):
    output_file = tmp_path / "result.pdf"
    mapped_data = {"顧客名": "テスト太郎", "金額": "¥150,000"}
    
    template = Template(
        name="テストテンプレート",
        output_format="pdf",
        output_filename_pattern="test.pdf",
        fields=[
            FieldMapping(source_key="name", output_label="顧客名", target_position=""),
            FieldMapping(source_key="price", output_label="金額", target_position=""),
        ]
    )
    
    try:
        import reportlab
    except ImportError:
        pytest.skip("reportlab がインストールされていません")
        
    exporter = ExporterFactory.create("pdf")
    exporter.export(mapped_data, template, output_file)
    
    assert output_file.exists()
