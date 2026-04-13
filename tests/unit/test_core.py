"""コア層のユニットテスト。"""

from pathlib import Path
from typing import Any

import pytest
import yaml

from app.core.exporter import (
    DocxExporter,
    ExporterFactory,
    PdfExporter,
    TxtExporter,
    XlsxExporter,
)
from app.core.ocr_engine import OCREngine
from app.core.template import (
    TemplateEngine,
    load_all_templates,
    load_template,
    load_template_set,
)
from app.models.ocr_result_model import OCRResult
from app.models.template_model import (
    FieldPlacement,
    Template,
    TemplateSet,
    TemplateSetEntry,
)


class MockOCREngine(OCREngine):
    """テスト用のモック OCR エンジン。"""

    def __init__(self, mock_data: dict[str, Any] | None = None):
        self._data = mock_data or {
            "invoice_no": "INV-001",
            "issue_date": "2026-04-11",
            "customer_name": "テスト商事",
            "total_amount": 12500,
        }

    def process(
        self,
        image_path: Path,
        extraction_prompt: str,
        response_schema: dict[str, Any],
        license_key: str,
    ) -> OCRResult:
        return OCRResult(
            source_image=image_path,
            extracted_data=self._data,
            processing_time_ms=100,
        )


def _make_template(output_format: str = "txt") -> Template:
    return Template(
        name="テストテンプレート",
        output_format=output_format,
        output_filename_pattern="test_output.{ext}".replace("{ext}", output_format),
        extraction_prompt="テスト用プロンプト",
        response_schema={"type": "object"},
        field_placements=[
            FieldPlacement(source_key="invoice_no", target="B2"),
            FieldPlacement(source_key="customer_name", target="B4"),
            FieldPlacement(source_key="total_amount", target="D25"),
        ],
        industry_preset="general",
    )


def _make_template_set() -> TemplateSet:
    return TemplateSet(
        name="テストセット",
        entries=[
            TemplateSetEntry(
                template_name="テストテンプレート",
                output_subfolder="invoices",
            ),
        ],
    )


class TestTemplateEngine:
    def test_apply_single(self, tmp_path: Path):
        engine = TemplateEngine()
        ocr = MockOCREngine()
        template = _make_template()
        image = tmp_path / "test.jpg"
        image.write_bytes(b"\xff\xd8\xff")

        result = engine.apply_single(ocr, image, template, "test-key")
        assert "B2" in result
        assert result["B2"] == "INV-001"
        assert result["__raw__"]["customer_name"] == "テスト商事"

    def test_apply_set(self, tmp_path: Path):
        engine = TemplateEngine()
        ocr = MockOCREngine()
        template = _make_template("txt")
        template_set = _make_template_set()
        image = tmp_path / "test.jpg"
        image.write_bytes(b"\xff\xd8\xff")
        output_root = tmp_path / "output"

        results = engine.apply_set(
            ocr_engine=ocr,
            image_path=image,
            template_set=template_set,
            templates_by_name={"テストテンプレート": template},
            output_root=output_root,
            license_key="test-key",
        )
        assert len(results) == 1
        assert results[0].status == "success"
        assert results[0].output_file is not None
        assert results[0].output_file.exists()

    def test_apply_set_missing_template(self, tmp_path: Path):
        engine = TemplateEngine()
        ocr = MockOCREngine()
        ts = TemplateSet(
            name="s",
            entries=[TemplateSetEntry(template_name="不在", output_subfolder="x")],
        )
        image = tmp_path / "test.jpg"
        image.write_bytes(b"\xff\xd8\xff")

        results = engine.apply_set(
            ocr_engine=ocr,
            image_path=image,
            template_set=ts,
            templates_by_name={},
            output_root=tmp_path,
            license_key="k",
        )
        assert results[0].status == "failed"
        assert "見つかりません" in results[0].error_message


class TestLoadTemplateYAML:
    def test_load_template(self, tmp_path: Path):
        yaml_content = {
            "name": "テスト",
            "output_format": "txt",
            "output_filename_pattern": "out.txt",
            "extraction_prompt": "テスト",
            "response_schema": {"type": "object"},
            "field_placements": [
                {"source_key": "a", "target": "B1"},
            ],
        }
        path = tmp_path / "test.yaml"
        path.write_text(yaml.dump(yaml_content, allow_unicode=True), encoding="utf-8")

        tmpl = load_template(path)
        assert tmpl.name == "テスト"
        assert len(tmpl.field_placements) == 1

    def test_load_template_set(self, tmp_path: Path):
        yaml_content = {
            "name": "セット",
            "entries": [
                {"template_name": "テスト", "output_subfolder": "out"},
            ],
        }
        path = tmp_path / "set.yaml"
        path.write_text(yaml.dump(yaml_content, allow_unicode=True), encoding="utf-8")

        ts = load_template_set(path)
        assert ts.name == "セット"

    def test_load_all_templates_uses_yaml_stem_as_key(self, tmp_path: Path):
        tpath = tmp_path / "my_invoice.yaml"
        tpath.write_text(
            yaml.dump(
                {
                    "name": "表示用名前",
                    "output_format": "txt",
                    "output_filename_pattern": "o.txt",
                    "extraction_prompt": "p",
                    "response_schema": {"type": "object"},
                },
                allow_unicode=True,
            ),
            encoding="utf-8",
        )
        loaded = load_all_templates([tmp_path])
        assert "my_invoice" in loaded
        assert loaded["my_invoice"].name == "表示用名前"


class TestExporters:
    def test_txt_exporter(self, tmp_path: Path):
        exporter = TxtExporter()
        template = _make_template("txt")
        data = {
            "__raw__": {"invoice_no": "INV-001", "total_amount": 12500},
            "B2": "INV-001",
        }
        out = tmp_path / "out.txt"
        exporter.export(data, template, out)
        assert out.read_bytes().startswith(b"\xef\xbb\xbf")
        content = out.read_text(encoding="utf-8-sig")
        assert "INV-001" in content

    def test_docx_exporter(self, tmp_path: Path):
        exporter = DocxExporter()
        template = _make_template("docx")
        data = {
            "__raw__": {"invoice_no": "INV-001"},
        }
        out = tmp_path / "out.docx"
        exporter.export(data, template, out)
        assert out.exists()

    def test_xlsx_exporter(self, tmp_path: Path):
        exporter = XlsxExporter()
        template = _make_template("xlsx")
        data = {
            "__raw__": {"invoice_no": "INV-001", "customer_name": "テスト", "total_amount": 100},
        }
        out = tmp_path / "out.xlsx"
        exporter.export(data, template, out)
        assert out.exists()
        from openpyxl import load_workbook

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None
        assert ws["B2"].value == "INV-001"
        assert ws["B4"].value == "テスト"
        assert ws["D25"].value == 100

    def test_exporter_factory(self):
        assert isinstance(ExporterFactory.create("txt"), TxtExporter)
        assert isinstance(ExporterFactory.create("docx"), DocxExporter)
        assert isinstance(ExporterFactory.create("xlsx"), XlsxExporter)
        assert isinstance(ExporterFactory.create("pdf"), PdfExporter)

    def test_exporter_factory_unsupported(self):
        from app.exceptions import ExportError
        with pytest.raises(ExportError):
            ExporterFactory.create("html")
